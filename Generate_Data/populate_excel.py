"""
Populate Tables.xlsx with data from generated CSVs and FK relationships.
Run from the repo root: python Generate_Data/populate_excel.py
"""
import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_DIR = os.path.join(REPO_ROOT, "Generate_Data")
EXCEL_PATH = os.path.join(REPO_ROOT, "Tables.xlsx")

# Maps sheet name → CSV filename
SHEET_CSV_MAP = {
    "Table-1":  "users.csv",
    "Table-2":  "categories.csv",
    "Table-3":  "products.csv",
    "Table-4":  "subscriptions.csv",
    "Table-5":  "payments.csv",
    "Table-6":  "reviews.csv",
    "Table-7":  "seller_payouts.csv",
    "Table-8":  "tags.csv",
    "Table-9":  "product_tags.csv",
}

FK_ROWS = [
    ("Products",        "Users",         "seller_id",       "user_id"),
    ("Products",        "Categories",    "category_id",     "category_id"),
    ("Product_Tags",    "Products",      "product_id",      "product_id"),
    ("Product_Tags",    "Tags",          "tag_id",          "tag_id"),
    ("Subscriptions",   "Users",         "buyer_id",        "user_id"),
    ("Subscriptions",   "Products",      "product_id",      "product_id"),
    ("Payments",        "Subscriptions", "subscription_id", "subscription_id"),
    ("Reviews",         "Products",      "product_id",      "product_id"),
    ("Reviews",         "Users",         "buyer_id",        "user_id"),
    ("Seller_Payouts",  "Users",         "seller_id",       "user_id"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col)].width = 20


def load_csv(filename):
    path = os.path.join(CSV_DIR, filename)
    if not os.path.exists(path):
        print(f"  WARNING: {filename} not found — skipping")
        return [], []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    headers = list(rows[0].keys()) if rows else []
    return headers, rows


def fill_sheet(ws, headers, rows):
    ws.delete_rows(1, ws.max_row)
    for col_idx, h in enumerate(headers, 1):
        ws.cell(1, col_idx, h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row_idx, col_idx, row.get(h, ""))
    style_header(ws, len(headers))
    print(f"  Wrote {len(rows)} data rows + header")


def fill_relation_table(ws):
    ws.delete_rows(1, ws.max_row)
    headers = ["REFERENCING TABLE", "REFERENCED TABLE", "FOREIGN KEY", "PRIMARY KEY"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(1, col_idx, h)
    for row_idx, (ref_table, ref_table2, fk, pk) in enumerate(FK_ROWS, 2):
        ws.cell(row_idx, 1, ref_table)
        ws.cell(row_idx, 2, ref_table2)
        ws.cell(row_idx, 3, fk)
        ws.cell(row_idx, 4, pk)
    style_header(ws, 4)
    print(f"  Wrote {len(FK_ROWS)} FK rows")


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    print("Filling RELATION TABLE sheet...")
    fill_relation_table(wb["RELATION TABLE"])

    for sheet_name, csv_file in SHEET_CSV_MAP.items():
        print(f"Filling {sheet_name} from {csv_file}...")
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet {sheet_name!r} not found in workbook — skipping")
            continue
        ws = wb[sheet_name]
        headers, rows = load_csv(csv_file)
        if headers:
            fill_sheet(ws, headers, rows)

    wb.save(EXCEL_PATH)
    print(f"\nSaved: {EXCEL_PATH}")

    # Verification
    wb2 = openpyxl.load_workbook(EXCEL_PATH)
    print("\nVerification:")
    for sheet_name in list(SHEET_CSV_MAP.keys()) + ["RELATION TABLE"]:
        if sheet_name in wb2.sheetnames:
            ws = wb2[sheet_name]
            print(f"  {sheet_name}: {ws.max_row - 1} data rows, {ws.max_column} cols")


if __name__ == "__main__":
    main()
